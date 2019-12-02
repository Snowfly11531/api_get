from openpyxl import Workbook,load_workbook,worksheet
from openpyxl.worksheet.worksheet import Worksheet
import os

class OperateExcel():
    def __init__(self):
        self.wb: Workbook
        self.ws: Worksheet
        self.__filePath: str

    def pastePathAndName(self,path,name):
        return path+"\\"+name+".xlsx"

    def __OpenOrCreateExcelWorkBook(self,filePath):
        if not os.path.exists(filePath):
            wb=Workbook()
            wb.save(filePath)
        self.wb=load_workbook(filePath)

    def OpenOrCreateExcelWorkBookByPathAndName(self,path,name):
        self.__filePath=self.pastePathAndName(path,name)
        self.__OpenOrCreateExcelWorkBook(self.__filePath)

    def setOrCreateActiveSheetByname(self,name):
        try:
            self.wb[name]
        except KeyError:
            self.wb.create_sheet(name)
        self.ws=self.wb[name]

    def insertRowAtEnd(self,rowlist):
        self.ws.append(rowlist)

    def insertColAtEnd(self,collist):
        columnCounts=len(list(self.ws.columns))
        for i in range(len(collist)):
            self.ws.cell(i+1,columnCounts+1,collist[i])

    def insertDirtList(self,DirtList:list):
        self.ws.delete_rows(1,len(list(self.ws.rows)))
        row=[]
        for key in list(dict(DirtList[0]).keys()):
            row.append(key)
        self.ws.append(row)
        for dirt in DirtList:
            row=[]
            for value in list(dict(dirt).values()):
                row.append(value)
            self.ws.append(row)

    def save(self):
        self.wb.save(self.__filePath)

    def test(self):
        wb=self.OpenOrCreateExcelWorkBookByPathAndName(path="D:",name="haha")
        ws=self.setOrCreateActiveSheetByname("haha")
        print(type(ws))
        self.insertRowAtEnd(rowlist=[1,2,3])
        self.insertColAtEnd(collist=[1,2,3,4])
        self.wb.save(self.pastePathAndName("D:", "haha"))

    def insertDataColByIDAtEnd(self,Dirtlist:list,id_col_name="ID号",data_col_name="data",alias_name=""):
        rows=list(self.ws.rows)
        index=0
        for cell in rows[0]:
            if cell.value==id_col_name:
                break
            index+=1
        print(index)



if __name__ == '__main__':
    pass